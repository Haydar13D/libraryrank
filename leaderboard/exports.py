from django.http import HttpResponse
from django.db.models import Count, Q
from django.utils import timezone
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.enums import TA_CENTER, TA_LEFT
import io

from .models import Member, Faculty, Book, BorrowRecord, Visit


def get_ranked_members(role, date_from, date_to):
    return Member.objects.filter(role=role, is_active=True).annotate(
        visit_total=Count('visits', filter=Q(
            visits__visited_at__date__gte=date_from,
            visits__visited_at__date__lte=date_to
        )),
        borrow_total=Count('borrow_records', filter=Q(
            borrow_records__borrowed_at__date__gte=date_from,
            borrow_records__borrowed_at__date__lte=date_to
        ))
    ).order_by('-visit_total')


# ─────────────── EXCEL ───────────────

def export_excel_response(date_from, date_to):
    wb = openpyxl.Workbook()

    # Styles
    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill_blue = PatternFill('solid', fgColor='1e3a5f')
    header_fill_green = PatternFill('solid', fgColor='1a4a2e')
    header_fill_orange = PatternFill('solid', fgColor='4a2e0a')
    header_fill_purple = PatternFill('solid', fgColor='2e1a4a')
    header_fill_dark = PatternFill('solid', fgColor='1a1a2e')
    center = Alignment(horizontal='center', vertical='center')
    thin = Border(
        left=Side(style='thin', color='dddddd'),
        right=Side(style='thin', color='dddddd'),
        top=Side(style='thin', color='dddddd'),
        bottom=Side(style='thin', color='dddddd'),
    )

    def make_sheet(ws, title, data_rows, headers, col_widths, fill):
        # Title row
        ws.merge_cells(f'A1:{get_column_letter(len(headers))}1')
        title_cell = ws.cell(row=1, column=1, value=title)
        title_cell.font = Font(name='Calibri', bold=True, size=14, color='FFFFFF')
        title_cell.fill = fill
        title_cell.alignment = center
        ws.row_dimensions[1].height = 30

        # Date range row
        ws.merge_cells(f'A2:{get_column_letter(len(headers))}2')
        date_cell = ws.cell(row=2, column=1, value=f'Period: {date_from} to {date_to}')
        date_cell.font = Font(name='Calibri', italic=True, size=10, color='888888')
        date_cell.alignment = center
        ws.row_dimensions[2].height = 18

        # Headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = fill
            cell.alignment = center
            cell.border = thin
        ws.row_dimensions[3].height = 22

        # Data
        for row_num, row_data in enumerate(data_rows, 4):
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col, value=val)
                cell.font = Font(name='Calibri', size=10)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin
                if row_num % 2 == 0:
                    cell.fill = PatternFill('solid', fgColor='f8f8ff')

        # Column widths
        for col, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

    # ── Students Sheet ──
    ws_s = wb.active
    ws_s.title = '🎓 Students'
    students = get_ranked_members('student', date_from, date_to)
    s_rows = [(i+1, m.member_id, m.name, m.faculty.name if m.faculty else '', m.year_enrolled or '', m.visit_total, m.borrow_total, m.streak_days)
              for i, m in enumerate(students)]
    make_sheet(ws_s, 'Student Leaderboard', s_rows,
               ['#', 'ID', 'Name', 'Faculty', 'Year', 'Visits', 'Books', 'Streak'],
               [5, 12, 25, 20, 8, 10, 10, 10], header_fill_blue)

    # ── Lecturers Sheet ──
    ws_l = wb.create_sheet('👨‍🏫 Lecturers')
    lecturers = get_ranked_members('lecturer', date_from, date_to)
    l_rows = [(i+1, m.member_id, m.name, m.faculty.name if m.faculty else '', m.title, m.visit_total, m.borrow_total, m.streak_days)
              for i, m in enumerate(lecturers)]
    make_sheet(ws_l, 'Lecturer Leaderboard', l_rows,
               ['#', 'ID', 'Name', 'Faculty', 'Title', 'Visits', 'Books', 'Streak'],
               [5, 12, 25, 20, 18, 10, 10, 10], header_fill_green)

    # ── Staff Sheet ──
    ws_st = wb.create_sheet('👔 Staff')
    staff = get_ranked_members('staff', date_from, date_to)
    st_rows = [(i+1, m.member_id, m.name, m.department, m.title or 'Staff', m.visit_total, m.borrow_total, m.streak_days)
               for i, m in enumerate(staff)]
    make_sheet(ws_st, 'Staff Leaderboard', st_rows,
               ['#', 'ID', 'Name', 'Department', 'Role', 'Visits', 'Books', 'Streak'],
               [5, 12, 25, 20, 18, 10, 10, 10], header_fill_orange)

    # ── Books Sheet ──
    ws_b = wb.create_sheet('📚 Books')
    books = Book.objects.annotate(
        borrow_total=Count('borrow_records', filter=Q(
            borrow_records__borrowed_at__date__gte=date_from,
            borrow_records__borrowed_at__date__lte=date_to
        ))
    ).order_by('-borrow_total')
    b_rows = [(i+1, bk.isbn, bk.title, bk.author, bk.category, bk.faculty.name if bk.faculty else '', bk.borrow_total)
              for i, bk in enumerate(books)]
    make_sheet(ws_b, 'Most Borrowed Books', b_rows,
               ['#', 'ISBN', 'Title', 'Author', 'Category', 'Faculty', 'Borrows'],
               [5, 16, 35, 25, 15, 18, 10], header_fill_purple)

    # ── Faculties Sheet ──
    ws_f = wb.create_sheet('🏛️ Faculties')
    from .models import Faculty as FacultyModel
    faculties = FacultyModel.objects.annotate(
        v=Count('members__visits', filter=Q(members__visits__visited_at__date__gte=date_from, members__visits__visited_at__date__lte=date_to)),
        bk=Count('members__borrow_records', filter=Q(members__borrow_records__borrowed_at__date__gte=date_from, members__borrow_records__borrowed_at__date__lte=date_to))
    ).order_by('-v')
    f_rows = [(i+1, f.name, f.code, f.v, f.bk) for i, f in enumerate(faculties)]
    make_sheet(ws_f, 'Faculty Rankings', f_rows,
               ['#', 'Faculty', 'Code', 'Visitors', 'Books Borrowed'],
               [5, 30, 10, 14, 18], header_fill_dark)

    # Save & return
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f'LibraryRank_{date_from}_to_{date_to}.xlsx'
    response = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ─────────────── PDF ───────────────

def export_pdf_response(date_from, date_to):
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            rightMargin=1.5*cm, leftMargin=1.5*cm,
                            topMargin=2*cm, bottomMargin=2*cm)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', fontName='Helvetica-Bold', fontSize=20,
                                  textColor=colors.HexColor('#f5c842'), alignment=TA_CENTER, spaceAfter=6)
    sub_style = ParagraphStyle('Sub', fontName='Helvetica', fontSize=10,
                                textColor=colors.HexColor('#888888'), alignment=TA_CENTER, spaceAfter=20)
    section_style = ParagraphStyle('Section', fontName='Helvetica-Bold', fontSize=13,
                                    textColor=colors.HexColor('#eef2ff'), spaceAfter=8, spaceBefore=16)

    DARK_BG = colors.HexColor('#0d0f14')
    SURFACE = colors.HexColor('#161923')
    GOLD = colors.HexColor('#f5c842')
    BLUE = colors.HexColor('#4da6ff')
    GREEN = colors.HexColor('#3de08a')
    ORANGE = colors.HexColor('#ff914d')
    WHITE = colors.HexColor('#eef2ff')
    MUTED = colors.HexColor('#8892aa')
    ROW_ALT = colors.HexColor('#1e2330')

    def make_table(headers, rows, header_color):
        table_data = [headers] + rows
        t = Table(table_data, repeatRows=1)
        style_cmds = [
            ('BACKGROUND', (0, 0), (-1, 0), header_color),
            ('TEXTCOLOR', (0, 0), (-1, 0), WHITE),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWHEIGHT', (0, 0), (-1, 0), 22),
            ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#252d3d')),
            ('BACKGROUND', (0, 1), (-1, -1), SURFACE),
            ('TEXTCOLOR', (0, 1), (-1, -1), WHITE),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ROWHEIGHT', (0, 1), (-1, -1), 18),
        ]
        for i in range(1, len(rows) + 1):
            if i % 2 == 0:
                style_cmds.append(('BACKGROUND', (0, i), (-1, i), ROW_ALT))
            # Highlight top 3
            if i == 1:
                style_cmds.append(('TEXTCOLOR', (0, i), (0, i), GOLD))
                style_cmds.append(('FONTNAME', (0, i), (0, i), 'Helvetica-Bold'))
            elif i == 2:
                style_cmds.append(('TEXTCOLOR', (0, i), (0, i), MUTED))
            elif i == 3:
                style_cmds.append(('TEXTCOLOR', (0, i), (0, i), ORANGE))

        t.setStyle(TableStyle(style_cmds))
        return t

    story = []

    # Cover
    story.append(Paragraph('📚 LibraryRank', title_style))
    story.append(Paragraph(f'University Library Leaderboard Report', sub_style))
    story.append(Paragraph(f'Period: {date_from}  →  {date_to}', sub_style))
    story.append(Spacer(1, 0.5*cm))

    # Students
    story.append(Paragraph('🎓 Student Rankings', section_style))
    students = get_ranked_members('student', date_from, date_to)
    s_rows = [[str(i+1), m.member_id, m.name, m.faculty.name if m.faculty else '', str(m.year_enrolled or ''), str(m.visit_total), str(m.borrow_total), f'{m.streak_days}d 🔥']
              for i, m in enumerate(students[:15])]
    story.append(make_table(['#', 'ID', 'Name', 'Faculty', 'Year', 'Visits', 'Books', 'Streak'], s_rows, BLUE))
    story.append(Spacer(1, 0.5*cm))

    # Lecturers
    story.append(Paragraph('👨‍🏫 Lecturer Rankings', section_style))
    lecturers = get_ranked_members('lecturer', date_from, date_to)
    l_rows = [[str(i+1), m.member_id, m.name, m.faculty.name if m.faculty else '', m.title, str(m.visit_total), str(m.borrow_total)]
              for i, m in enumerate(lecturers[:10])]
    story.append(make_table(['#', 'ID', 'Name', 'Faculty', 'Title', 'Visits', 'Books'], l_rows, GREEN))
    story.append(Spacer(1, 0.5*cm))

    # Staff
    story.append(Paragraph('👔 Staff Rankings', section_style))
    staff = get_ranked_members('staff', date_from, date_to)
    st_rows = [[str(i+1), m.member_id, m.name, m.department, m.title or 'Staff', str(m.visit_total)]
               for i, m in enumerate(staff)]
    story.append(make_table(['#', 'ID', 'Name', 'Department', 'Role', 'Visits'], st_rows, ORANGE))
    story.append(Spacer(1, 0.5*cm))

    # Books
    story.append(Paragraph('📚 Most Borrowed Books', section_style))
    books = Book.objects.annotate(
        borrow_total=Count('borrow_records', filter=Q(
            borrow_records__borrowed_at__date__gte=date_from,
            borrow_records__borrowed_at__date__lte=date_to
        ))
    ).order_by('-borrow_total')[:10]
    b_rows = [[str(i+1), bk.title[:45], bk.author[:30], bk.category, str(bk.borrow_total)]
              for i, bk in enumerate(books)]
    story.append(make_table(['#', 'Title', 'Author', 'Category', 'Borrows'], b_rows, GOLD))

    doc.build(story)
    buf.seek(0)
    filename = f'LibraryRank_{date_from}_to_{date_to}.pdf'
    response = HttpResponse(buf.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
